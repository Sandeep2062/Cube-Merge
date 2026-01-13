import openpyxl
from openpyxl.drawing.image import Image as XLImage
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import os
import winsound
from copy import deepcopy
import shutil
import json
import sys

"""
╔══════════════════════════════════════════════════════════════════╗
║                    CUBE DATA PROCESSOR                            ║
║                                                                   ║
║  Developer: Sandeep (https://github.com/Sandeep2062)            ║
║  Repository: https://github.com/Sandeep2062/Cube-Data-Processor ║
║                                                                   ║
║  © 2026 Sandeep - All Rights Reserved                           ║
╚══════════════════════════════════════════════════════════════════╝
"""

# Settings file location (same folder as EXE)
def get_settings_path():
    """Get settings file path in same folder as script/EXE"""
    if getattr(sys, 'frozen', False):
        # Running as EXE
        base_path = os.path.dirname(sys.executable)
    else:
        # Running as script
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(base_path, "cube_settings.json")

SETTINGS_FILE = get_settings_path()

# Load saved settings
def load_settings():
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r') as f:
                settings = json.load(f)
                # Convert grade_files list to filenames only
                if "grade_files" in settings:
                    grade_files_data = settings["grade_files"]
                else:
                    grade_files_data = []
                return settings, grade_files_data
    except:
        pass
    return {"output_path": "", "calendar_path": ""}, []

# Save settings (remembers grade files and paths except office)
def save_settings(grade_file_list, output, calendar):
    try:
        settings = {
            "grade_files": grade_file_list,  # Save full paths
            "output_path": output,
            "calendar_path": calendar
        }
        with open(SETTINGS_FILE, 'w') as f:
            json.dump(settings, f, indent=2)
    except Exception as e:
        print(f"Could not save settings: {e}")

# Smart grade extraction - handles M20, M15, and Mortar_1_4 format
def extract_grade(filename):
    name = os.path.basename(filename).split('.')[0].upper()
    
    # Check if it's mortar format (contains underscore followed by numbers)
    # E.g., MORTAR_1_4 → 1:4
    if "MORTAR" in name and "_" in name:
        parts = name.split("_")
        # Get the ratio part after MORTAR_
        if len(parts) >= 3:  # MORTAR_1_4
            ratio = f"{parts[-2]}:{parts[-1]}"
            return ratio
    
    # For regular grades (M20, M15, etc.) - just clean up
    name = name.replace("_", "").replace("-", "")
    return name.strip()

# FILLED ROW CHECKER
def get_last_row(ws):
    row = 2
    while True:
        if ws.cell(row=row, column=2).value in (None, ""):
            return row - 1
        row += 1

# SAFE WORKBOOK LOADING
def load_workbook_safe(filepath):
    try:
        wb = openpyxl.load_workbook(filepath, keep_vba=False, data_only=False, keep_links=False)
        return wb
    except:
        wb = openpyxl.load_workbook(filepath)
        return wb

# LOAD CALENDAR DATA
def load_calendar_data(calendar_file, log):
    """Load calendar dates from Excel file"""
    try:
        if not calendar_file or not os.path.exists(calendar_file):
            log("⚠ No calendar file selected")
            return None
        
        wb = load_workbook_safe(calendar_file)
        ws = wb.active
        
        calendar_dict = {}
        row = 2  # Start from row 2
        
        while True:
            casting_date = ws.cell(row=row, column=1).value  # Column A
            if not casting_date:
                break
            
            date_7 = ws.cell(row=row, column=2).value   # Column B (7 days)
            date_28 = ws.cell(row=row, column=3).value  # Column C (28 days)
            
            # Store as string for matching
            if casting_date:
                date_str = str(casting_date).strip()
                calendar_dict[date_str] = {
                    "7_days": str(date_7).strip() if date_7 else "",
                    "28_days": str(date_28).strip() if date_28 else ""
                }
            
            row += 1
        
        wb.close()
        log(f"✓ Calendar loaded: {len(calendar_dict)} dates")
        return calendar_dict
        
    except Exception as e:
        log(f"✖ Calendar load error: {e}")
        return None

# PROCESS WITH GRADE AND/OR DATE
def process_combined(grade_files, office_file, output_folder, calendar_file, mode, log):
    try:
        log(f"\n{'='*60}")
        log(f"PROCESSING MODE: {mode.upper().replace('_', ' ')}")
        log(f"{'='*60}")
        
        # Load calendar if date mode
        calendar_data = None
        if mode in ["date_only", "both"]:
            calendar_data = load_calendar_data(calendar_file, log)
            if not calendar_data:
                log("✖ Cannot proceed without calendar file")
                return 0
        
        # Create output filename (NO TIMESTAMP)
        base = os.path.basename(office_file).split(".")[0]
        outname = f"{base}_Processed.xlsx"
        outpath = os.path.join(output_folder, outname)
        
        # Copy template to preserve images
        shutil.copy2(office_file, outpath)
        
        # Load and modify the copy
        office_wb = load_workbook_safe(outpath)
        
        total_copy_count = 0
        
        # GRADE PROCESSING
        if mode in ["grade_only", "both"] and grade_files:
            log(f"\n--- GRADE PROCESSING ---")
            
            for grade_file in grade_files:
                grade_wb = load_workbook_safe(grade_file)
                grade_ws = grade_wb.active
                grade_name = extract_grade(grade_file)
                
                log(f"\nProcessing: {os.path.basename(grade_file)}")
                log(f"Looking for grade: {grade_name}")
                
                last_row = get_last_row(grade_ws)
                log(f"Data rows: {last_row - 1}")
                
                # Find matching sheets - normalize both sides for comparison
                matching_sheets = []
                for sheet_name in office_wb.sheetnames:
                    ws = office_wb[sheet_name]
                    b12_value = ws["B12"].value
                    if b12_value:
                        # Normalize B12 value
                        b12 = str(b12_value).replace(" ", "").upper()
                        
                        # For comparison, also normalize grade_name
                        grade_normalized = grade_name.replace(" ", "").upper()
                        
                        if b12 == grade_normalized:
                            matching_sheets.append(sheet_name)
                            log(f"  ✓ Matched sheet: {sheet_name} (B12={b12})")
                
                log(f"Total matching sheets: {len(matching_sheets)}")
                
                if len(matching_sheets) == 0:
                    log(f"⚠ No sheets found with B12='{grade_name}'")
                    grade_wb.close()
                    continue
                
                sheet_index = 0
                
                # Copy grade data
                for r in range(2, last_row + 1):
                    if sheet_index >= len(matching_sheets):
                        log(f"⚠ More data rows than available sheets")
                        break
                    
                    current_sheet_name = matching_sheets[sheet_index]
                    ws = office_wb[current_sheet_name]
                    
                    weight_values = [grade_ws.cell(row=r, column=c).value for c in range(2, 8)]
                    strength_values = [grade_ws.cell(row=r, column=c).value for c in range(9, 15)]
                    
                    for i, v in enumerate(weight_values):
                        ws.cell(row=25, column=3 + i, value=v)
                    for i, v in enumerate(strength_values):
                        ws.cell(row=27, column=3 + i, value=v)
                    
                    total_copy_count += 1
                    log(f"  ✓ Row {r} → {current_sheet_name}")
                    sheet_index += 1
                
                grade_wb.close()
        
        # DATE PROCESSING
        if mode in ["date_only", "both"] and calendar_data:
            log(f"\n--- DATE PROCESSING ---")
            
            updated_count = 0
            
            for sheet_name in office_wb.sheetnames:
                ws = office_wb[sheet_name]
                
                # Read casting date from C17
                casting_date_cell = ws["C17"].value
                if not casting_date_cell:
                    continue
                
                casting_date = str(casting_date_cell).strip()
                
                # Look up in calendar
                if casting_date in calendar_data:
                    date_7 = calendar_data[casting_date]["7_days"]
                    date_28 = calendar_data[casting_date]["28_days"]
                    
                    # Write 7-day date to C18
                    if date_7:
                        ws["C18"] = date_7
                    
                    # Write 28-day date to F18
                    if date_28:
                        ws["F18"] = date_28
                    
                    updated_count += 1
                    log(f"✓ {sheet_name}: {casting_date} → 7d:{date_7}, 28d:{date_28}")
                else:
                    log(f"⚠ Date not in calendar: {casting_date} ({sheet_name})")
            
            log(f"\nSheets updated: {updated_count}")
        
        # Save combined file
        office_wb.save(outpath)
        office_wb.close()
        
        log(f"\n{'='*60}")
        log(f"✓✓✓ SAVED: {outpath}")
        log(f"{'='*60}")
        
        return total_copy_count
        
    except Exception as e:
        log(f"✖ ERROR: {e}")
        import traceback
        log(traceback.format_exc())
        return 0

# ------------- GUI LOGIC -------------

def run_processing():
    # Validate inputs based on mode
    mode = mode_var.get()
    
    if mode in ["grade_only", "both"]:
        if not grade_files:
            messagebox.showerror("Error", "Please select grade files for grade processing.")
            return
    
    if mode in ["date_only", "both"]:
        if not calendar_path.get():
            messagebox.showerror("Error", "Please select calendar file for date processing.")
            return
    
    if not office_path.get():
        messagebox.showerror("Error", "Select office format file.")
        return

    if not output_path.get():
        messagebox.showerror("Error", "Select output folder.")
        return

    # Save settings (grade files, output path, calendar - NOT office)
    save_settings(grade_files, output_path.get(), calendar_path.get())

    log_box.delete("1.0", "end")
    
    progress["value"] = 50
    root.update_idletasks()
    
    total = process_combined(
        grade_files,
        office_path.get(),
        output_path.get(),
        calendar_path.get(),
        mode,
        log=lambda m: log_box.insert(tk.END, m + "\n")
    )
    
    progress["value"] = 100
    
    winsound.MessageBeep()
    messagebox.showinfo("✓ Completed", f"Processing Complete!\n\nTotal Operations: {total}")

def add_grades():
    files = filedialog.askopenfilenames(filetypes=[("Excel Files", "*.xlsx")])
    for f in files:
        if f not in grade_files:
            grade_files.append(f)
            grade_listbox.insert(tk.END, os.path.basename(f))

def clear_grades():
    grade_files.clear()
    grade_listbox.delete(0, tk.END)

def pick_office():
    path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if path:
        office_path.set(path)

def pick_calendar():
    path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if path:
        calendar_path.set(path)

def pick_output_folder():
    folder = filedialog.askdirectory()
    if folder:
        output_path.set(folder)

def update_mode_ui():
    """Show/hide sections based on mode"""
    mode = mode_var.get()
    
    if mode == "grade_only":
        grade_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10), before=office_frame)
        calendar_frame.pack_forget()
    elif mode == "date_only":
        grade_frame.pack_forget()
        calendar_frame.pack(fill=tk.X, pady=(0, 10), before=office_frame)
    else:  # both
        grade_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10), before=office_frame)
        calendar_frame.pack(fill=tk.X, pady=(0, 10), before=office_frame)

# --------------------------- V6.2 ----------------------------

root = tk.Tk()
root.title("Cube Data Processor")
root.geometry("1000x900")
root.minsize(950, 850)

try:
    root.iconbitmap("icon.ico")
except:
    pass

# Load saved settings
settings, saved_grade_files = load_settings()

grade_files = []
office_path = tk.StringVar()
output_path = tk.StringVar(value=settings.get("output_path", ""))
calendar_path = tk.StringVar(value=settings.get("calendar_path", ""))
mode_var = tk.StringVar(value="both")

# Enhanced Color Scheme for Beautiful UI
BG_PRIMARY = "#0d1117"          # GitHub dark background
BG_SECONDARY = "#161b22"        # Slightly lighter
BG_CARD = "#21262d"             # Card background
BG_HOVER = "#30363d"            # Hover state
BG_INPUT = "#0d1117"            # Input field background
TEXT_PRIMARY = "#f0f6fc"        # Crisp white text
TEXT_SECONDARY = "#8b949e"      # Muted text
ACCENT_PRIMARY = "#58a6ff"      # GitHub blue
ACCENT_SECONDARY = "#1f6feb"    # Darker blue
ACCENT_SUCCESS = "#3fb950"      # GitHub green
ACCENT_WARNING = "#d29922"      # GitHub yellow
BORDER_COLOR = "#30363d"        # Border color
PROGRESS_COLOR = "#3fb950"      # Progress bar color

# Set root background
root.configure(bg=BG_PRIMARY)

# Styles
style = ttk.Style()
style.theme_use('clam')

# Configure button styles
style.configure('Primary.TButton', 
                padding=12, 
                relief="flat", 
                background=ACCENT_PRIMARY, 
                foreground=TEXT_PRIMARY, 
                font=("Segoe UI", 10, "bold"), 
                borderwidth=0)
style.map('Primary.TButton', 
          background=[('active', ACCENT_SECONDARY)])

style.configure('Success.TButton', 
                padding=16, 
                background=ACCENT_SUCCESS, 
                foreground=BG_PRIMARY, 
                font=("Segoe UI", 12, "bold"), 
                borderwidth=0)
style.map('Success.TButton', 
          background=[('active', '#2ea043')])

style.configure('Card.TFrame', 
                background=BG_CARD, 
                relief="flat", 
                borderwidth=1, 
                focusthickness=0)

style.configure("Horizontal.TProgressbar", 
                background=PROGRESS_COLOR, 
                troughcolor=BG_INPUT, 
                borderwidth=0, 
                thickness=10)

# Enhanced Header with Modern Design
header_frame = tk.Frame(root, bg=BG_PRIMARY, height=120)
header_frame.pack(fill=tk.X)
header_frame.pack_propagate(False)

# Create a top bar with gradient effect
top_bar = tk.Frame(header_frame, bg=ACCENT_PRIMARY, height=4)
top_bar.pack(fill=tk.X)

# Main header container with subtle gradient background
header_bg = tk.Frame(header_frame, bg=BG_CARD)
header_bg.pack(fill=tk.BOTH, expand=True, padx=30, pady=10)

# Left side - Logo and Title
left_section = tk.Frame(header_bg, bg=BG_CARD)
left_section.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Logo with elegant frame
logo_frame = tk.Frame(left_section, bg=BG_CARD, highlightbackground=ACCENT_PRIMARY, highlightthickness=2)
logo_frame.pack(side=tk.LEFT, padx=(0, 20))

try:
    from PIL import Image, ImageTk
    logo_img = Image.open("logo.png")
    logo_img = logo_img.resize((70, 70), Image.Resampling.LANCZOS)
    logo_photo = ImageTk.PhotoImage(logo_img)
    logo_label = tk.Label(logo_frame, image=logo_photo, bg=BG_CARD)
    logo_label.image = logo_photo
    logo_label.pack(padx=5, pady=5)
except:
    # Fallback to emoji with styling
    logo_label = tk.Label(logo_frame, text="🔷", font=("Segoe UI", 48), 
                         bg=BG_CARD, fg=ACCENT_PRIMARY)
    logo_label.pack(padx=5, pady=5)

# Title section
title_section = tk.Frame(left_section, bg=BG_CARD)
title_section.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Main title with better styling
title_label = tk.Label(title_section, text="CUBE DATA PROCESSOR", 
                       font=("Segoe UI", 32, "bold"), bg=BG_CARD, fg=TEXT_PRIMARY, 
                       anchor="w")
title_label.pack(anchor="w", pady=(5, 0))

# Subtitle with icon
subtitle_container = tk.Frame(title_section, bg=BG_CARD)
subtitle_container.pack(anchor="w", pady=(2, 0))

subtitle_icon = tk.Label(subtitle_container, text="⚡", font=("Segoe UI", 12), 
                        bg=BG_CARD, fg=ACCENT_WARNING)
subtitle_icon.pack(side=tk.LEFT, padx=(0, 5))

subtitle_label = tk.Label(subtitle_container, text="Professional Edition v2.0", 
                         font=("Segoe UI", 12, "italic"), bg=BG_CARD, fg=TEXT_SECONDARY)
subtitle_label.pack(side=tk.LEFT)

# Status indicator
status_frame = tk.Frame(title_section, bg=BG_CARD)
status_frame.pack(anchor="w", pady=(5, 0))

status_dot = tk.Label(status_frame, text="●", font=("Segoe UI", 10), 
                     bg=BG_CARD, fg=ACCENT_SUCCESS)
status_dot.pack(side=tk.LEFT, padx=(0, 5))

status_label = tk.Label(status_frame, text="Ready", 
                       font=("Segoe UI", 10), bg=BG_CARD, fg=ACCENT_SUCCESS)
status_label.pack(side=tk.LEFT)

# Right side - Developer Info
right_section = tk.Frame(header_bg, bg=BG_CARD)
right_section.pack(side=tk.RIGHT, padx=(20, 0))

# Developer card with elegant styling
dev_card = tk.Frame(right_section, bg=BG_SECONDARY, relief=tk.RAISED, bd=1)
dev_card.pack(pady=10)

# Developer header
dev_header = tk.Frame(dev_card, bg=ACCENT_PRIMARY)
dev_header.pack(fill=tk.X, padx=2, pady=2)

dev_title = tk.Label(dev_header, text="DEVELOPER", 
                    font=("Segoe UI", 9, "bold"), bg=ACCENT_PRIMARY, fg="white")
dev_title.pack(pady=5)

# Developer info
dev_info_frame = tk.Frame(dev_card, bg=BG_SECONDARY)
dev_info_frame.pack(padx=15, pady=10)

# Developer name with icon
dev_name_frame = tk.Frame(dev_info_frame, bg=BG_SECONDARY)
dev_name_frame.pack(anchor="w", pady=(0, 5))

dev_icon = tk.Label(dev_name_frame, text="👨‍💻", font=("Segoe UI", 16), 
                   bg=BG_SECONDARY, fg=TEXT_PRIMARY)
dev_icon.pack(side=tk.LEFT, padx=(0, 8))

dev_name_label = tk.Label(dev_name_frame, text="SANDEEP", 
                         font=("Segoe UI", 16, "bold"), bg=BG_SECONDARY, fg=TEXT_PRIMARY)
dev_name_label.pack(side=tk.LEFT)

# GitHub link with button styling
github_frame = tk.Frame(dev_info_frame, bg=BG_SECONDARY)
github_frame.pack(anchor="w", pady=(5, 0))

github_btn = tk.Button(github_frame, text="🔗 GitHub Profile", 
                      font=("Segoe UI", 9, "bold"), bg=BG_CARD, fg=ACCENT_PRIMARY,
                      activebackground=BG_HOVER, activeforeground=ACCENT_SECONDARY,
                      relief=tk.FLAT, cursor="hand2", padx=10, pady=5, bd=0,
                      command=lambda: os.system("start https://github.com/Sandeep2062/Cube-Data-Processor"))
github_btn.pack()

# Add hover effect to GitHub button
def github_on_enter(e):
    github_btn.configure(bg=BG_HOVER)

def github_on_leave(e):
    github_btn.configure(bg=BG_CARD)

github_btn.bind("<Enter>", github_on_enter)
github_btn.bind("<Leave>", github_on_leave)

# Version info at bottom
version_frame = tk.Frame(dev_card, bg=BG_SECONDARY)
version_frame.pack(fill=tk.X, padx=2, pady=(0, 2))

version_label = tk.Label(version_frame, text="Version 2.0.0 | Build 2026.1", 
                        font=("Segoe UI", 8), bg=BG_SECONDARY, fg=TEXT_SECONDARY)
version_label.pack(pady=3)

# Main Container with Scrollbar
main_canvas = tk.Canvas(root, bg=BG_PRIMARY, highlightthickness=0)
main_scrollbar = tk.Scrollbar(root, orient="vertical", command=main_canvas.yview, bg=BG_SECONDARY, troughcolor=BG_PRIMARY)
scrollable_frame = tk.Frame(main_canvas, bg=BG_PRIMARY)

main_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
main_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
canvas_frame = main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

def configure_scroll_region(event):
    main_canvas.configure(scrollregion=main_canvas.bbox("all"))

def configure_canvas_width(event):
    canvas_width = event.width
    main_canvas.itemconfig(canvas_frame, width=canvas_width)

scrollable_frame.bind("<Configure>", configure_scroll_region)
main_canvas.bind("<Configure>", configure_canvas_width)

# Mouse wheel scrolling
def on_mousewheel(event):
    main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")

main_canvas.bind_all("<MouseWheel>", on_mousewheel)

# Main container frame (inside scrollable area)
main_container = tk.Frame(scrollable_frame, bg=BG_PRIMARY)
main_container.pack(fill=tk.BOTH, expand=True, padx=30, pady=25)

# Processing Mode Selection
mode_selection_frame = tk.LabelFrame(main_container, text="  ⚙️ PROCESSING MODE  ", 
                                    font=("Segoe UI", 12, "bold"), bg=BG_CARD, 
                                    fg=TEXT_PRIMARY, bd=0, relief=tk.FLAT, 
                                    padx=25, pady=20, highlightbackground=BORDER_COLOR, 
                                    highlightthickness=1)
mode_selection_frame.pack(fill=tk.X, pady=(0, 15))

# Custom Radio Buttons with better styling
radio_frame = tk.Frame(mode_selection_frame, bg=BG_CARD)
radio_frame.pack(fill=tk.X, pady=5)

def create_radio_button(parent, text, value, row):
    frame = tk.Frame(parent, bg=BG_CARD)
    frame.grid(row=row, column=0, sticky="w", pady=8)
    
    var = tk.BooleanVar()
    radio = tk.Radiobutton(frame, text=text, variable=mode_var, value=value, 
                          font=("Segoe UI", 11), bg=BG_CARD, fg=TEXT_PRIMARY, 
                          activebackground=BG_CARD, activeforeground=TEXT_PRIMARY, 
                          selectcolor=BG_INPUT, command=update_mode_ui, cursor="hand2")
    radio.pack(side=tk.LEFT, padx=5)
    
    # Add hover effect
    def on_enter(e):
        frame.configure(bg=BG_HOVER)
        radio.configure(bg=BG_HOVER)
    
    def on_leave(e):
        frame.configure(bg=BG_CARD)
        radio.configure(bg=BG_CARD)
    
    frame.bind("<Enter>", on_enter)
    frame.bind("<Leave>", on_leave)
    radio.bind("<Enter>", on_enter)
    radio.bind("<Leave>", on_leave)
    
    return frame

grade_radio = create_radio_button(radio_frame, "📊 Grade Only", "grade_only", 0)
date_radio = create_radio_button(radio_frame, "📅 Date Only", "date_only", 1)
both_radio = create_radio_button(radio_frame, "🔄 Both (Grade + Date)", "both", 2)

# Grade Files Section
grade_frame = tk.LabelFrame(main_container, text="  📁 GRADE FILES  ", 
                            font=("Segoe UI", 12, "bold"), bg=BG_CARD, 
                            fg=TEXT_PRIMARY, bd=0, relief=tk.FLAT, 
                            padx=25, pady=20, highlightbackground=BORDER_COLOR,
                            highlightthickness=1)

btn_frame = tk.Frame(grade_frame, bg=BG_CARD)
btn_frame.pack(fill=tk.X, pady=(0, 15))

add_btn = tk.Button(btn_frame, text="➕ Add Files", command=add_grades, 
                   bg=ACCENT_PRIMARY, fg=TEXT_PRIMARY, font=("Segoe UI", 10, "bold"),
                   activebackground=ACCENT_SECONDARY, activeforeground=TEXT_PRIMARY,
                   relief=tk.FLAT, cursor="hand2", padx=15, pady=8, bd=0)
add_btn.pack(side=tk.LEFT, padx=5)

clear_btn = tk.Button(btn_frame, text="🗑️ Clear", command=clear_grades, 
                     bg="#da3633", fg=TEXT_PRIMARY, font=("Segoe UI", 10, "bold"),
                     activebackground="#b91c1c", activeforeground=TEXT_PRIMARY,
                     relief=tk.FLAT, cursor="hand2", padx=15, pady=8, bd=0)
clear_btn.pack(side=tk.LEFT, padx=5)

# Add hover effects to buttons
def add_hover_effect(button, bg_color, hover_color):
    def on_enter(e):
        button.configure(bg=hover_color)
    
    def on_leave(e):
        button.configure(bg=bg_color)
    
    button.bind("<Enter>", on_enter)
    button.bind("<Leave>", on_leave)

add_hover_effect(add_btn, ACCENT_PRIMARY, ACCENT_SECONDARY)
add_hover_effect(clear_btn, "#da3633", "#b91c1c")

grade_listbox = tk.Listbox(grade_frame, height=4, font=("Consolas", 10), 
                           bg=BG_INPUT, fg=TEXT_PRIMARY, relief=tk.FLAT, bd=0, 
                           highlightthickness=1, highlightbackground=BORDER_COLOR,
                           selectbackground=ACCENT_PRIMARY, selectforeground="white")
grade_listbox.pack(fill=tk.BOTH, expand=True)

# Load saved grade files
for gf in saved_grade_files:
    if os.path.exists(gf):
        grade_files.append(gf)
        grade_listbox.insert(tk.END, os.path.basename(gf))

# Calendar File Section
calendar_frame = tk.LabelFrame(main_container, text="  📅 CALENDAR FILE  ", 
                              font=("Segoe UI", 12, "bold"), bg=BG_CARD, 
                              fg=TEXT_PRIMARY, bd=0, relief=tk.FLAT, 
                              padx=25, pady=20, highlightbackground=BORDER_COLOR,
                              highlightthickness=1)

calendar_btn = tk.Button(calendar_frame, text="📂 Select Calendar", command=pick_calendar, 
                         bg=ACCENT_PRIMARY, fg=TEXT_PRIMARY, font=("Segoe UI", 10, "bold"),
                         activebackground=ACCENT_SECONDARY, activeforeground=TEXT_PRIMARY,
                         relief=tk.FLAT, cursor="hand2", padx=15, pady=8, bd=0)
calendar_btn.pack(anchor=tk.W, pady=(0, 15))
add_hover_effect(calendar_btn, ACCENT_PRIMARY, ACCENT_SECONDARY)

calendar_entry = tk.Entry(calendar_frame, textvariable=calendar_path, font=("Segoe UI", 10),
                         bg=BG_INPUT, fg=TEXT_PRIMARY, relief=tk.FLAT, bd=0, 
                         insertbackground=TEXT_PRIMARY, highlightthickness=1,
                         highlightbackground=BORDER_COLOR)
calendar_entry.pack(fill=tk.X, ipady=12, padx=2)

# Office File Section
office_frame = tk.LabelFrame(main_container, text="  📄 OFFICE FORMAT FILE  ", 
                            font=("Segoe UI", 12, "bold"), bg=BG_CARD, 
                            fg=TEXT_PRIMARY, bd=0, relief=tk.FLAT, 
                            padx=25, pady=20, highlightbackground=BORDER_COLOR,
                            highlightthickness=1)
office_frame.pack(fill=tk.X, pady=(0, 15))

office_btn = tk.Button(office_frame, text="📂 Select File", command=pick_office, 
                      bg=ACCENT_PRIMARY, fg=TEXT_PRIMARY, font=("Segoe UI", 10, "bold"),
                      activebackground=ACCENT_SECONDARY, activeforeground=TEXT_PRIMARY,
                      relief=tk.FLAT, cursor="hand2", padx=15, pady=8, bd=0)
office_btn.pack(anchor=tk.W, pady=(0, 15))
add_hover_effect(office_btn, ACCENT_PRIMARY, ACCENT_SECONDARY)

office_entry = tk.Entry(office_frame, textvariable=office_path, font=("Segoe UI", 10),
                       bg=BG_INPUT, fg=TEXT_PRIMARY, relief=tk.FLAT, bd=0, 
                       insertbackground=TEXT_PRIMARY, highlightthickness=1,
                       highlightbackground=BORDER_COLOR)
office_entry.pack(fill=tk.X, ipady=12, padx=2)

# Output Folder Section
output_frame = tk.LabelFrame(main_container, text="  💾 OUTPUT FOLDER  ", 
                            font=("Segoe UI", 12, "bold"), bg=BG_CARD, 
                            fg=TEXT_PRIMARY, bd=0, relief=tk.FLAT, 
                            padx=25, pady=20, highlightbackground=BORDER_COLOR,
                            highlightthickness=1)
output_frame.pack(fill=tk.X, pady=(0, 15))

output_btn = tk.Button(output_frame, text="📂 Select Folder", command=pick_output_folder, 
                      bg=ACCENT_PRIMARY, fg=TEXT_PRIMARY, font=("Segoe UI", 10, "bold"),
                      activebackground=ACCENT_SECONDARY, activeforeground=TEXT_PRIMARY,
                      relief=tk.FLAT, cursor="hand2", padx=15, pady=8, bd=0)
output_btn.pack(anchor=tk.W, pady=(0, 15))
add_hover_effect(output_btn, ACCENT_PRIMARY, ACCENT_SECONDARY)

output_entry = tk.Entry(output_frame, textvariable=output_path, font=("Segoe UI", 10),
                       bg=BG_INPUT, fg=TEXT_PRIMARY, relief=tk.FLAT, bd=0, 
                       insertbackground=TEXT_PRIMARY, highlightthickness=1,
                       highlightbackground=BORDER_COLOR)
output_entry.pack(fill=tk.X, ipady=12, padx=2)

# Start Button with enhanced styling
start_btn_frame = tk.Frame(main_container, bg=BG_PRIMARY)
start_btn_frame.pack(pady=(25, 20))

# Add glow effect behind button
glow_frame = tk.Frame(start_btn_frame, bg=BG_PRIMARY, highlightbackground=ACCENT_SUCCESS, highlightthickness=2)
glow_frame.pack(padx=10, pady=10)

start_btn = tk.Button(glow_frame, text="▶️  START PROCESSING", command=run_processing,
                     font=("Segoe UI", 14, "bold"), bg=ACCENT_SUCCESS, fg=BG_PRIMARY,
                     activebackground="#2ea043", relief=tk.FLAT, cursor="hand2",
                     padx=40, pady=18, bd=0)
start_btn.pack()

def start_btn_on_enter(e):
    start_btn.configure(bg="#2ea043")
    glow_frame.configure(highlightbackground="#2ea043")

def start_btn_on_leave(e):
    start_btn.configure(bg=ACCENT_SUCCESS)
    glow_frame.configure(highlightbackground=ACCENT_SUCCESS)

start_btn.bind("<Enter>", start_btn_on_enter)
start_btn.bind("<Leave>", start_btn_on_leave)

# Progress Bar
progress_frame = tk.Frame(main_container, bg=BG_PRIMARY)
progress_frame.pack(fill=tk.X, pady=(0, 15))

progress = ttk.Progressbar(progress_frame, length=700, mode="determinate", 
                          style="Horizontal.TProgressbar")
progress.pack()

# Log Section
log_frame = tk.LabelFrame(main_container, text="  📋 PROCESSING LOG  ", 
                         font=("Segoe UI", 12, "bold"), bg=BG_CARD, 
                         fg=TEXT_PRIMARY, bd=0, relief=tk.FLAT, 
                         padx=25, pady=20, highlightbackground=BORDER_COLOR,
                         highlightthickness=1)
log_frame.pack(fill=tk.BOTH, expand=True)

log_scrollbar = tk.Scrollbar(log_frame, bg=BG_INPUT)
log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

log_box = tk.Text(log_frame, height=10, font=("Consolas", 10), bg=BG_INPUT, 
                 fg=ACCENT_SUCCESS, relief=tk.FLAT, bd=0, wrap=tk.WORD, 
                 yscrollcommand=log_scrollbar.set, insertbackground=TEXT_PRIMARY)
log_box.pack(fill=tk.BOTH, expand=True)
log_scrollbar.config(command=log_box.yview)

# Enhanced Footer
footer = tk.Frame(root, bg=BG_SECONDARY, height=50)
footer.pack(fill=tk.X, side=tk.BOTTOM)
footer.pack_propagate(False)

# Footer content
footer_content = tk.Frame(footer, bg=BG_SECONDARY)
footer_content.pack(expand=True)

# Left side - copyright
copyright_label = tk.Label(footer_content, text="© 2026 Sandeep", 
                          font=("Segoe UI", 9), bg=BG_SECONDARY, fg=TEXT_SECONDARY)
copyright_label.pack(side=tk.LEFT, padx=20)

# Center - status
footer_status = tk.Label(footer_content, text="● System Ready", 
                        font=("Segoe UI", 9), bg=BG_SECONDARY, fg=ACCENT_SUCCESS)
footer_status.pack(side=tk.LEFT, expand=True)

# Right side - links
links_frame = tk.Frame(footer_content, bg=BG_SECONDARY)
links_frame.pack(side=tk.RIGHT, padx=20)

github_link = tk.Label(links_frame, text="GitHub", 
                      font=("Segoe UI", 9), bg=BG_SECONDARY, fg=ACCENT_PRIMARY, 
                      cursor="hand2")
github_link.pack(side=tk.LEFT, padx=10)
github_link.bind("<Button-1>", lambda e: os.system("start https://github.com/Sandeep2062/Cube-Data-Processor"))

separator = tk.Label(links_frame, text="|", font=("Segoe UI", 9), 
                    bg=BG_SECONDARY, fg=TEXT_SECONDARY)
separator.pack(side=tk.LEFT)

docs_link = tk.Label(links_frame, text="Documentation", 
                    font=("Segoe UI", 9), bg=BG_SECONDARY, fg=ACCENT_PRIMARY, 
                    cursor="hand2")
docs_link.pack(side=tk.LEFT, padx=10)

# Initialize UI
update_mode_ui()

root.mainloop()